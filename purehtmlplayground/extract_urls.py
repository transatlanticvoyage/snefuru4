#!/usr/bin/env python3
import os
import openpyxl
import re

folder_path = "/Users/kylecampbell/Documents/repos/localrepo-snefuru4/purehtmlplayground/social profiles - main guy - normal deliveries"
output_file = os.path.join(folder_path, "agg1.txt")

all_urls = []

# Pattern to match URLs
url_pattern = re.compile(r'https?://[^\s]+')

# Process each Excel file
for filename in os.listdir(folder_path):
    if filename.endswith('.xlsx') and not filename.startswith('~$'):
        file_path = os.path.join(folder_path, filename)
        print(f"Processing: {filename}")
        
        try:
            # Open the Excel file
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            # Process each sheet
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                # Find columns with headers containing "Profiles URLs" or "Live Backlinks"
                header_row = None
                target_columns = []
                
                # Search for header row (usually in first 10 rows)
                for row_idx in range(1, min(11, sheet.max_row + 1)):
                    row_has_header = False
                    for col_idx in range(1, sheet.max_column + 1):
                        cell_value = sheet.cell(row=row_idx, column=col_idx).value
                        if cell_value:
                            cell_str = str(cell_value).lower()
                            if 'profiles url' in cell_str or 'live backlink' in cell_str or 'profile url' in cell_str:
                                header_row = row_idx
                                target_columns.append(col_idx)
                                row_has_header = True
                                print(f"  Found header '{cell_value}' at row {row_idx}, col {col_idx}")
                    if row_has_header:
                        break
                
                # If we found relevant columns, extract URLs from those columns
                if target_columns and header_row:
                    # Start from the row after the header
                    for row_idx in range(header_row + 1, sheet.max_row + 1):
                        for col_idx in target_columns:
                            cell_value = sheet.cell(row=row_idx, column=col_idx).value
                            if cell_value:
                                cell_str = str(cell_value).strip()
                                # Check if it's a URL
                                if cell_str.startswith(('http://', 'https://')):
                                    all_urls.append(cell_str)
            
            wb.close()
            
        except Exception as e:
            print(f"Error processing {filename}: {e}")

# Remove duplicates and sort
unique_urls = sorted(set(all_urls))

# Write to output file
with open(output_file, 'w') as f:
    for url in unique_urls:
        f.write(url + '\n')

print(f"\nTotal unique URLs found: {len(unique_urls)}")
print(f"URLs saved to: {output_file}")