#!/usr/bin/env python3
import os
import openpyxl

folder_path = "/Users/kylecampbell/Documents/repos/localrepo-snefuru4/purehtmlplayground/social profiles - main guy - bulk deliveries"

total_urls_found = 0
total_files = 0

# Check each Excel file
for filename in os.listdir(folder_path):
    if filename.endswith('.xlsx') and not filename.startswith('~$'):
        file_path = os.path.join(folder_path, filename)
        total_files += 1
        
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            # Check each sheet
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                # Find the URL column
                header_row = None
                url_column = None
                
                for row_idx in range(1, min(21, sheet.max_row + 1)):
                    for col_idx in range(1, sheet.max_column + 1):
                        cell_value = sheet.cell(row=row_idx, column=col_idx).value
                        if cell_value and 'Profiles URLs' in str(cell_value):
                            header_row = row_idx
                            url_column = col_idx
                            break
                    if header_row:
                        break
                
                if header_row and url_column:
                    # Count URLs in this sheet
                    sheet_urls = 0
                    for row_idx in range(header_row + 1, sheet.max_row + 1):
                        cell_value = sheet.cell(row=row_idx, column=url_column).value
                        if cell_value:
                            cell_str = str(cell_value).strip()
                            if cell_str.startswith(('http://', 'https://')):
                                sheet_urls += 1
                    
                    print(f"{filename} - {sheet_name}: {sheet_urls} URLs")
                    total_urls_found += sheet_urls
                else:
                    print(f"{filename} - {sheet_name}: No URL column found")
            
            wb.close()
            
        except Exception as e:
            print(f"Error processing {filename}: {e}")

print(f"\n=== SUMMARY ===")
print(f"Total files processed: {total_files}")
print(f"Total URLs found in spreadsheets: {total_urls_found}")

# Compare with extracted file
extracted_file = os.path.join(folder_path, "agg1 - created by claude code.txt")
if os.path.exists(extracted_file):
    with open(extracted_file, 'r') as f:
        extracted_urls = len([line for line in f if line.strip()])
    print(f"Total URLs in extracted file: {extracted_urls}")
    print(f"URLs missing from extraction: {total_urls_found - extracted_urls}")
else:
    print("Extracted file not found")