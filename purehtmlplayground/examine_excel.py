#!/usr/bin/env python3
import os
import openpyxl

# Let's examine one file to understand its structure
folder_path = "/Users/kylecampbell/Documents/repos/localrepo-snefuru4/purehtmlplayground/social profiles - main guy - normal deliveries"
sample_file = "17636378451604830689_HQ_Profile_Backlinks_Report_For___Decatur_Electrician_Group.xlsx"
file_path = os.path.join(folder_path, sample_file)

print(f"Examining: {sample_file}\n")

wb = openpyxl.load_workbook(file_path, data_only=True)

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"\nSheet: {sheet_name}")
    print(f"Max rows: {sheet.max_row}, Max columns: {sheet.max_column}")
    
    # Print first 15 rows to see the structure
    print("\nFirst 15 rows:")
    for row_idx in range(1, min(16, sheet.max_row + 1)):
        row_values = []
        for col_idx in range(1, min(10, sheet.max_column + 1)):  # First 10 columns
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            if cell_value:
                row_values.append(str(cell_value)[:50])  # Truncate long values
        if row_values:
            print(f"Row {row_idx}: {row_values}")

wb.close()