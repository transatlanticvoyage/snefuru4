#!/usr/bin/env python3
import os
import openpyxl
from collections import defaultdict

normal_folder = "/Users/kylecampbell/Documents/repos/localrepo-snefuru4/purehtmlplayground/social profiles - main guy - normal deliveries"
extracted_file = os.path.join(normal_folder, "agg1 - created by claude code.txt")

print("=== NORMAL DELIVERY FOLDER ANALYSIS REPORT ===\n")

# Check each file for URLs
total_urls_in_files = 0
file_details = []

print("1. URL EXTRACTION FROM SPREADSHEETS:")
files = [f for f in os.listdir(normal_folder) if f.endswith('.xlsx') and not f.startswith('~$')]

for filename in files:
    file_path = os.path.join(normal_folder, filename)
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            # Find the URL column
            header_row = None
            url_column = None
            
            for row_idx in range(1, min(21, sheet.max_row + 1)):
                for col_idx in range(1, sheet.max_column + 1):
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value
                    if cell_value and ('Profiles URLs' in str(cell_value) or 'Live Backlinks' in str(cell_value)):
                        header_row = row_idx
                        url_column = col_idx
                        break
                if header_row:
                    break
            
            if header_row and url_column:
                # Count URLs in this sheet
                sheet_urls = 0
                for row_idx in range(header_row + 1, sheet.max_row + 1):
                    cell_value = sheet.cell(row=row_idx, column=url_column).value
                    if cell_value:
                        cell_str = str(cell_value).strip()
                        if cell_str.startswith(('http://', 'https://')):
                            sheet_urls += 1
                
                business_name = sheet_name
                file_details.append((filename, business_name, sheet_urls))
                total_urls_in_files += sheet_urls
                print(f"   • {business_name}: {sheet_urls} URLs")
        
        wb.close()
        
    except Exception as e:
        print(f"   ERROR processing {filename}: {e}")

print(f"\nTotal URLs found in spreadsheets: {total_urls_in_files}")

# Check extracted file
if os.path.exists(extracted_file):
    with open(extracted_file, 'r') as f:
        extracted_urls = [line.strip() for line in f if line.strip()]
    
    print(f"Total URLs in extracted file: {len(extracted_urls)}")
    
    missing_count = total_urls_in_files - len(extracted_urls)
    if missing_count > 0:
        print(f"⚠️  URLs MISSING from extraction: {missing_count}")
    else:
        print("✅ ALL URLs successfully extracted - no URLs skipped!")

print("\n" + "="*70)

# Analyze business breakdown from extracted URLs
print("\n2. BREAKDOWN BY TARGET PROPERTY/BUSINESS:\n")

if os.path.exists(extracted_file):
    # Define business patterns more comprehensively
    business_patterns = {
        'ForgePinnacle Roofing (Kirkland)': ['forgepinnacle', 'roofingskirklandwa', 'roofingcontractorskirklandwa', 'kirklandroof'],
        'Richardson Pest Control': ['richardsonpestcontrol', 'pestcontrolrichardsontx', 'pestcontrolrich', 'pestcontrolprostx'],
        'Decatur Electrician Group': ['decaturelectrician', 'electriciandecatural', 'electricianal', 'decatur01'],
        'Naperville Pest Control': ['napervillepestcontrol', 'pestcontrolguysil', 'pestcontrolguys', 'napervillepe'],
        'Lexington Water Damage': ['lexingtonwaterdamage', 'waterdamagelexingtonky', 'waterdamageky'],
        'ShellBeam Roofing (Alpharetta)': ['shellbeam', 'alpharettaroofing', 'shellbeofalph', 'alpharettaroofi']
    }
    
    business_counts = defaultdict(list)
    unmatched_urls = []
    
    for url in extracted_urls:
        matched = False
        for business, patterns in business_patterns.items():
            if any(pattern.lower() in url.lower() for pattern in patterns):
                business_counts[business].append(url)
                matched = True
                break
        if not matched:
            unmatched_urls.append(url)
    
    # Sort by count
    sorted_businesses = sorted(business_counts.items(), key=lambda x: len(x[1]), reverse=True)
    
    for business, urls in sorted_businesses:
        print(f"   • {business}: {len(urls)} URLs")
    
    accounted_urls = sum(len(urls) for urls in business_counts.values())
    print(f"\n   Total URLs accounted for: {accounted_urls}")
    
    if unmatched_urls:
        print(f"   ⚠️ Unmatched URLs: {len(unmatched_urls)}")
        print("   First few unmatched URLs:")
        for url in unmatched_urls[:5]:
            print(f"     - {url}")

print("\n" + "="*70)

print("\n3. DETAILED FILE BREAKDOWN:\n")
for filename, business, count in file_details:
    clean_filename = filename.replace('_HQ_Profile_Backlinks_Report_For_', '').replace('.xlsx', '')
    print(f"   • {business}")
    print(f"     File: {clean_filename}")
    print(f"     URLs extracted: {count}")
    print()

print("4. FINAL SUMMARY:")
print(f"   • Excel files processed: {len(files)}")
print(f"   • Total businesses: {len(file_details)}")
print(f"   • Total URLs in spreadsheets: {total_urls_in_files}")
if os.path.exists(extracted_file):
    print(f"   • URLs successfully extracted: {len(extracted_urls)}")
    extraction_rate = (len(extracted_urls)/total_urls_in_files*100) if total_urls_in_files > 0 else 0
    print(f"   • Extraction success rate: {extraction_rate:.1f}%")
    if len(extracted_urls) == total_urls_in_files:
        print(f"   ✅ PERFECT EXTRACTION - No URLs were skipped!")
    else:
        print(f"   ⚠️ Missing URLs: {total_urls_in_files - len(extracted_urls)}")
else:
    print("   • No extraction file found")

print("\n" + "="*70)