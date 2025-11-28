#!/usr/bin/env python3
import os
import openpyxl

folder_path = "/Users/kylecampbell/Documents/repos/localrepo-snefuru4/purehtmlplayground/social profiles - main guy - bulk deliveries"

# Let's examine a few files to see if there are any URLs we missed
sample_files = ["02ce5a162c99a593a99d45f2e2316454.xlsx", "16a4590347b689f0151c26bf19186313.xlsx"]

for filename in sample_files[:2]:  # Check first 2 files
    file_path = os.path.join(folder_path, filename)
    print(f"\n=== Examining: {filename} ===")
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(f"\nSheet: {sheet_name}")
            print(f"Max rows: {sheet.max_row}, Max columns: {sheet.max_column}")
            
            # Find the header row
            header_row = None
            url_column = None
            
            for row_idx in range(1, min(21, sheet.max_row + 1)):
                for col_idx in range(1, sheet.max_column + 1):
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value
                    if cell_value and 'Profiles URLs' in str(cell_value):
                        header_row = row_idx
                        url_column = col_idx
                        print(f"Found header at row {row_idx}, column {col_idx}")
                        break
                if header_row:
                    break
            
            if header_row and url_column:
                # Count total non-empty cells in the URL column
                url_count = 0
                empty_count = 0
                non_url_count = 0
                
                for row_idx in range(header_row + 1, sheet.max_row + 1):
                    cell_value = sheet.cell(row=row_idx, column=url_column).value
                    if cell_value:
                        cell_str = str(cell_value).strip()
                        if cell_str.startswith(('http://', 'https://')):
                            url_count += 1
                        else:
                            non_url_count += 1
                            if non_url_count <= 5:  # Show first 5 non-URL values
                                print(f"  Non-URL value at row {row_idx}: {cell_str}")
                    else:
                        empty_count += 1
                
                print(f"URLs found: {url_count}")
                print(f"Non-URL values: {non_url_count}")
                print(f"Empty cells: {empty_count}")
                print(f"Total data rows: {sheet.max_row - header_row}")
            
        wb.close()
        
    except Exception as e:
        print(f"Error processing {filename}: {e}")

# Also check if there are any cells with URLs in other columns
print(f"\n=== Checking for URLs in other columns ===")
try:
    filename = "02ce5a162c99a593a99d45f2e2316454.xlsx"
    file_path = os.path.join(folder_path, filename)
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    sheet = wb[wb.sheetnames[0]]
    
    # Check all columns for URLs
    for col_idx in range(1, min(10, sheet.max_column + 1)):
        url_found_in_col = False
        for row_idx in range(1, min(50, sheet.max_row + 1)):
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            if cell_value:
                cell_str = str(cell_value).strip()
                if cell_str.startswith(('http://', 'https://')) and 'Profiles URLs' not in cell_str:
                    if not url_found_in_col:
                        print(f"Column {col_idx}: Found URLs")
                        url_found_in_col = True
                    if row_idx <= 20:  # Show first few URLs
                        print(f"  Row {row_idx}: {cell_str}")
    
    wb.close()
    
except Exception as e:
    print(f"Error in column check: {e}")