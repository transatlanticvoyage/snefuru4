#!/usr/bin/env python3
import os
import openpyxl
from collections import defaultdict

normal_folder = "/Users/kylecampbell/Documents/repos/localrepo-snefuru4/purehtmlplayground/social profiles - main guy - normal deliveries"
extracted_file = os.path.join(normal_folder, "agg1.txt")

print("=== NORMAL DELIVERY FOLDER ANALYSIS REPORT ===\n")

# First, check what files are in the folder
print("1. FILES IN FOLDER:")
files = [f for f in os.listdir(normal_folder) if f.endswith('.xlsx') and not f.startswith('~$')]
for i, filename in enumerate(files, 1):
    print(f"   {i}. {filename}")

print(f"\nTotal Excel files: {len(files)}\n")

# Check each file for URLs
total_urls_in_files = 0
file_details = []

print("2. URL COUNT PER FILE:")
for filename in files:
    file_path = os.path.join(normal_folder, filename)
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            # Find the URL column
            header_row = None
            url_column = None
            
            for row_idx in range(1, min(21, sheet.max_row + 1)):
                for col_idx in range(1, sheet.max_column + 1):
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value
                    if cell_value and ('Profiles URLs' in str(cell_value) or 'Live Backlinks' in str(cell_value)):
                        header_row = row_idx
                        url_column = col_idx
                        break
                if header_row:
                    break
            
            if header_row and url_column:
                # Count URLs in this sheet
                sheet_urls = 0
                for row_idx in range(header_row + 1, sheet.max_row + 1):
                    cell_value = sheet.cell(row=row_idx, column=url_column).value
                    if cell_value:
                        cell_str = str(cell_value).strip()
                        if cell_str.startswith(('http://', 'https://')):
                            sheet_urls += 1
                
                # Extract business name from filename or sheet
                business_name = sheet_name
                file_details.append((filename, business_name, sheet_urls))
                total_urls_in_files += sheet_urls
                print(f"   {filename} ({sheet_name}): {sheet_urls} URLs")
        
        wb.close()
        
    except Exception as e:
        print(f"   ERROR processing {filename}: {e}")

print(f"\nTotal URLs found in all spreadsheets: {total_urls_in_files}")

# Check extracted file
if os.path.exists(extracted_file):
    with open(extracted_file, 'r') as f:
        extracted_urls = [line.strip() for line in f if line.strip()]
    
    print(f"Total URLs in extracted file: {len(extracted_urls)}")
    
    missing_count = total_urls_in_files - len(extracted_urls)
    if missing_count > 0:
        print(f"⚠️  URLs MISSING from extraction: {missing_count}")
    else:
        print("✅ No URLs missing from extraction")
else:
    print("❌ Extracted file 'agg1.txt' not found!")

print("\n" + "="*60)

# Analyze business breakdown from extracted URLs
print("\n3. BREAKDOWN BY TARGET PROPERTY/BUSINESS:")

if os.path.exists(extracted_file):
    business_patterns = {
        'Richardson Pest Control': ['richardsonpestcontrol', 'pestcontrolrichardsontx'],
        'Naperville Pest Control': ['napervillepestcontrol', 'pestcontrolguys'],
        'ForgePinnacle Roofing': ['forgepinnacle', 'roofingskirklandwa', 'roofingcontractorskirklandwa'],
        'Lexington Water Damage': ['lexingtonwaterdamage', 'waterdamagelexingtonky'],
        'ShellBeam Roofing': ['shellbeam', 'alpharettaroofing'],
        'Decatur Electrician': ['decaturelectrician', 'electriciandecatural'],
        'Orem Electrician': ['oremelectrician', 'electricianoremutah'],
        'Chesapeake Porta Potty': ['chesapeakeppc', 'chesapeakeportapotty']
    }
    
    business_counts = defaultdict(int)
    
    for url in extracted_urls:
        for business, patterns in business_patterns.items():
            if any(pattern.lower() in url.lower() for pattern in patterns):
                business_counts[business] += 1
                break
    
    # Sort by count
    sorted_businesses = sorted(business_counts.items(), key=lambda x: x[1], reverse=True)
    
    for business, count in sorted_businesses:
        print(f"   {business}: {count} URLs")
    
    accounted_urls = sum(business_counts.values())
    print(f"\nTotal URLs accounted for: {accounted_urls}")
    print(f"Unaccounted URLs: {len(extracted_urls) - accounted_urls}")

print("\n" + "="*60)

print("\n4. SUMMARY:")
print(f"   • Files processed: {len(files)}")
print(f"   • Total URLs in spreadsheets: {total_urls_in_files}")
if os.path.exists(extracted_file):
    print(f"   • URLs successfully extracted: {len(extracted_urls)}")
    print(f"   • Extraction success rate: {(len(extracted_urls)/total_urls_in_files*100):.1f}%")
else:
    print("   • No extraction file found")

print("\n" + "="*60)