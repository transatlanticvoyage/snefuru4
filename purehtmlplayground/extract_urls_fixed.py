#!/usr/bin/env python3
import os
import openpyxl
import re

folder_path = "/Users/kylecampbell/Documents/repos/localrepo-snefuru4/purehtmlplayground/social profiles - main guy - normal deliveries"
output_file = os.path.join(folder_path, "agg1.txt")

all_urls = []

# Process each Excel file
for filename in os.listdir(folder_path):
    if filename.endswith('.xlsx') and not filename.startswith('~$'):
        file_path = os.path.join(folder_path, filename)
        print(f"Processing: {filename}")
        
        try:
            # Open the Excel file
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            # Process each sheet
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                # Search for the header row containing "Profiles URLs" or "Live Backlinks"
                header_row = None
                url_column = None
                
                # Search in first 20 rows for the header
                for row_idx in range(1, min(21, sheet.max_row + 1)):
                    for col_idx in range(1, sheet.max_column + 1):
                        cell_value = sheet.cell(row=row_idx, column=col_idx).value
                        if cell_value:
                            cell_str = str(cell_value)
                            # Check for the specific header
                            if 'Profiles URLs' in cell_str or 'Live Backlinks' in cell_str:
                                header_row = row_idx
                                url_column = col_idx
                                print(f"  Found header '{cell_str}' at row {row_idx}, column {col_idx}")
                                break
                    if header_row:
                        break
                
                # If we found the header, extract URLs from that column
                if header_row and url_column:
                    # Start from the row after the header
                    for row_idx in range(header_row + 1, sheet.max_row + 1):
                        cell_value = sheet.cell(row=row_idx, column=url_column).value
                        if cell_value:
                            cell_str = str(cell_value).strip()
                            # Check if it's a URL (starts with http:// or https://)
                            if cell_str.startswith(('http://', 'https://')):
                                # Clean the URL (sometimes Excel adds trailing characters)
                                url = cell_str.split()[0] if ' ' in cell_str else cell_str
                                all_urls.append(url)
                                
            wb.close()
            
        except Exception as e:
            print(f"Error processing {filename}: {e}")

# Remove duplicates while preserving order
seen = set()
unique_urls = []
for url in all_urls:
    if url not in seen:
        seen.add(url)
        unique_urls.append(url)

# Write to output file
with open(output_file, 'w') as f:
    for url in unique_urls:
        f.write(url + '\n')

print(f"\nTotal URLs found: {len(all_urls)}")
print(f"Total unique URLs: {len(unique_urls)}")
print(f"URLs saved to: {output_file}")